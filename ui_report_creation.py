# -*- coding: utf-8 -*-

################################################################################
## Form generated from reading UI file 'report_creation.ui'
##
## Created by: Qt User Interface Compiler version 6.5.2
##
## WARNING! All changes made in this file will be lost when recompiling UI file!
################################################################################

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtSql import QSqlDatabase, QSqlQuery
from PySide6.QtWidgets import (QApplication, QHeaderView, QLabel, QPushButton,
                               QSizePolicy, QTableWidget, QTableWidgetItem, QWidget, QMessageBox, QDialog)


import pandas as pd
import os
import shutil

from time import time, sleep

from sqlalchemy import text, create_engine
from openpyxl import load_workbook

import settings
import ui_animal_add
import ui_animal_edit
from openpyxl.styles import Alignment
from settings import DB_PATH, EXCEL_TEMPLATE_PATH, MAIN_REPORT_PAGE, EXCEL_HEADER_ROWS, SAVE_DIR, WRAP_COLUMNS, NAMES_TXT_PATH
from db_utils import db_get_city_name, db_get_owner, db_get_settlement, vet_db_connection


class Ui_Form(object):
    def transher_household_pk(self,household_pk):
        self.household_pk=household_pk
    def __init__(self):
        self.vet_db_connection = create_engine(f'sqlite:///{DB_PATH}').connect()
    def create_report(self, selected_items):
        settlement      = selected_items['settlement']
        city            = selected_items['city']
        city_name       = db_get_city_name(city)
        settlement_name = db_get_settlement(settlement)

        report_entries = pd.read_sql(text(
            f'''SELECT household.owner, city.name, household.address, report_entries.specie, report_entries.count, report_entries.data_from_administration, report_entries.prevous_count, report_entries.is_conditions_good FROM report_entries
                --присоединяем данные о хозяйстве (владелец и его адрес в рамках города)
                INNER JOIN household
                ON household.pk = report_entries.household

                --присоединяем данные о городе (для хозяйства, уточняем адрес) 
                INNER JOIN city
                ON city.pk = household.belongs_to_city
                WHERE city.pk = {city}
                     
            '''
        ), self.vet_db_connection)
        print(report_entries)
        # ----------------------------------------------------------------------------------дефолтные прелбразования данныхз ради некривого форматирования--------------------------------------------------------------------------------------------------------------------------
        report_entries.insert(0, 'position', report_entries['owner'].astype('category').cat.codes + 1)
        report_entries = report_entries.sort_values('owner')
        report_entries.loc[report_entries[['owner']].duplicated(keep='first'), ['owner', 'address', 'name', 'position']] = ''
        report_entries['address'] = report_entries.apply(lambda x: f"{x['name']}, {x['address']}" if x['address'] else "", axis=1)
        report_entries = report_entries.drop(columns='name')
        print(report_entries)
        
        # ----------------------------------------------------------------------------------создаем диреткории если нет и копируем шаблон, затем заполняем его--------------------------------------------------------------------------------------------------------------------------
        os.makedirs(SAVE_DIR, exist_ok=True)
        filename       = os.path.join(SAVE_DIR, f'Отчёт {int(time())}.xlsx')
        shutil.copy(EXCEL_TEMPLATE_PATH, filename)
        current_report = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
        
        # заполнение данными
        with open(NAMES_TXT_PATH, 'r', encoding='utf-8') as names_file:
            names = [line.strip() for line in names_file.readlines()]
        names = {'VET_CEO': names[0], 'VET_DOC': names[1], 'VET_DEP': names[2], 'CITY': city_name, 'SETTLEMENT': settlement_name}

        page = current_report.sheets[MAIN_REPORT_PAGE]
        message = QMessageBox()
        message.setText("loh")
        message.show()
        sleep(1.5)


        def fill_placeholders(names, cell):
            if isinstance(cell.value, str):
                for placeholder, name in names.items():
                    cell.value = cell.value.replace(placeholder, name)
        for row in range(1, page.max_row + 1):
            fill_placeholders(names, page.cell(row, 1))

        page.insert_rows(EXCEL_HEADER_ROWS + 1, amount=report_entries.shape[0])
        for row in range(report_entries.shape[0]):
            for col, col_name in enumerate(report_entries.columns):
                cell = page.cell(row + EXCEL_HEADER_ROWS + 1, col + 1)
                cell.value = report_entries.iloc[row, col]
                # перенос по словам
                if col_name in WRAP_COLUMNS:
                    cell.alignment = cell.alignment.copy(wrapText=True)

        current_report.close()
        
        
    def delete_animal(self):
        DB_PATH = settings.DB_PATH  # bezvremennoe reshenie
        VetDbConnnection = QSqlDatabase.addDatabase("QSQLITE")
        VetDbConnnection.setDatabaseName(DB_PATH)
        VetDbConnnection.open()
        VetTableQuery = QSqlQuery()
        vet_query_str="""DELETE FROM report_entries WHERE pk =""" + self.tableWidget.item(self.tableWidget.currentRow(), 0).text()
        VetTableQuery.prepare(vet_query_str)

        #VetTableQuery.bindValue(":deleted_pk", self.tableWidget.item(self.tableWidget.currentRow(), 0))

        uspeh = VetTableQuery.exec()
        print("USPEH UDALENIA &", uspeh, VetTableQuery.lastQuery())
        VetDbConnnection.close()
    def open_animals_add(self):
        self.window = QWidget()
        self.ui = ui_animal_add.Ui_Form()
        self.ui.transfer_animal_add_data(self.household_pk)
        self.ui.setupUi(self.window)
        
        print ("В юи репорт креайшн передан пк хозяйства; ", self.household_pk)
        self.window.show()
    def open_animals_edit(self):
        self.window = QWidget()
        self.ui = ui_animal_edit.Ui_Form()
        print(self.tableWidget.item(self.tableWidget.currentRow(), 4).text())
        print(type(self.tableWidget.item(self.tableWidget.currentRow(), 4)))
        self.ui.transfer_animal_add_data(
            self.household_pk,
            self.tableWidget.item(self.tableWidget.currentRow(), 0).text(),
            self.tableWidget.item(self.tableWidget.currentRow(), 1).text(),
            self.tableWidget.item(self.tableWidget.currentRow(), 2).text(),
            self.tableWidget.item(self.tableWidget.currentRow(), 3).text(),
            self.tableWidget.item(self.tableWidget.currentRow(), 4).text(),
            self.tableWidget.item(self.tableWidget.currentRow(), 5).text()
        )
        self.ui.setupUi(self.window)
        self.window.show()

    def setupUi(self, Form, city, address, owner, household_pk=1):
        self.household_pk = household_pk
        if not Form.objectName():
            Form.setObjectName(u"Form")
        Form.resize(758, 527)
        print(household_pk)
        font = QFont()
        font.setPointSize(9)
        Form.setFont(font)
        icon = QIcon(QIcon.fromTheme(u"address-book-new"))
        Form.setWindowIcon(icon)
        self.label = QLabel(Form)
        self.label.setObjectName(u"label")
        self.label.setGeometry(QRect(10, 50, 81, 41))
        font1 = QFont()
        font1.setPointSize(14)
        self.label.setFont(font1)
        self.label_2 = QLabel(Form)
        self.label_2.setObjectName(u"label_2")
        self.label_2.setGeometry(QRect(10, 90, 91, 41))
        self.label_2.setFont(font1)
        self.label_3 = QLabel(Form)
        self.label_3.setObjectName(u"label_3")
        self.label_3.setGeometry(QRect(10, 10, 81, 41))
        self.label_3.setFont(font1)
        self.placeForCity = QLabel(Form)
        self.placeForCity.setObjectName(u"placeForCity")
        self.placeForCity.setGeometry(QRect(90, 10, 251, 41))
        self.placeForCity.setFont(font1)
        self.placeForAddress = QLabel(Form)
        self.placeForAddress.setObjectName(u"placeForAddress")
        self.placeForAddress.setGeometry(QRect(100, 50, 251, 41))
        self.placeForAddress.setFont(font1)
        self.placeForOwner = QLabel(Form)
        self.placeForOwner.setObjectName(u"placeForOwner")
        self.placeForOwner.setGeometry(QRect(110, 90, 251, 41))
        self.placeForOwner.setFont(font1)
        self.tableWidget = QTableWidget(Form)
        if (self.tableWidget.columnCount() < 6):
            self.tableWidget.setColumnCount(6)
        __qtablewidgetitem = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, __qtablewidgetitem)
        __qtablewidgetitem1 = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, __qtablewidgetitem1)
        __qtablewidgetitem2 = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, __qtablewidgetitem2)
        __qtablewidgetitem3 = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(3, __qtablewidgetitem3)
        __qtablewidgetitem4 = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(4, __qtablewidgetitem4)
        __qtablewidgetitem5 = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(5, __qtablewidgetitem5)

        #self.tableWidget.setHorizontalHeaderItem(2, __qtablewidgetitem4) # а вот оно бля где ссылалось (внизу хуцйня)
        self.tableWidget.setObjectName(u"tableWidget")
        self.tableWidget.setGeometry(QRect(20, 160, 701, 281))
        self.label_4 = QLabel(Form)
        self.label_4.setObjectName(u"label_4")
        self.label_4.setGeometry(QRect(10, 120, 191, 51))
        font2 = QFont()
        font2.setPointSize(12)
        self.label_4.setFont(font2)
        self.documentAddAnimal = QPushButton(Form, clicked=lambda:self.open_animals_add())
        self.documentAddAnimal.setObjectName(u"documentAddAnimal")
        self.documentAddAnimal.setGeometry(QRect(20, 470, 141, 41))
        # self.createHouseholdReport = QPushButton(Form, clicked=lambda:self.create_report())
        # self.createHouseholdReport.setObjectName(u"createHouseholdReport")
        # self.createHouseholdReport.setGeometry(QRect(540, 460, 191, 51))
        # self.createHouseholdReport.setToolTipDuration(2)
        self.documentDeleteAnimal = QPushButton(Form, clicked=lambda:self.delete_animal())
        self.documentDeleteAnimal.setObjectName(u"documentDeleteAnimal")
        self.documentDeleteAnimal.setGeometry(QRect(340, 470, 141, 41))
        self.documentAddAnimal_2 = QPushButton(Form, clicked=lambda:self.open_animals_edit())#i forgor to rename it, it is actually edit button
        self.documentAddAnimal_2.setObjectName(u"documentAddAnimal_2")
        self.documentAddAnimal_2.setGeometry(QRect(170, 470, 161, 41))

        self.placeForCity.setText(QCoreApplication.translate("Form", city, None))
        self.placeForAddress.setText(QCoreApplication.translate("Form", address, None))
        self.placeForOwner.setText(QCoreApplication.translate("Form", owner, None))
        # -----------------animals_table------------------------
        pandas_SQL_query = f'SELECT pk, specie, count, data_from_administration, prevous_count, is_conditions_good FROM report_entries WHERE household ={household_pk}'
        print("ZAPROS NA TABLICY S ZHIVOTNIMI", pandas_SQL_query)
        data_for_table = pd.read_sql(text(pandas_SQL_query), vet_db_connection).astype(str)
        print("ZAPROS VOZVRASHAET",data_for_table )
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setRowCount(len(data_for_table))

        for col_num in range(len(data_for_table.columns)):
            for row_num in range(0, len(data_for_table)):
                self.tableWidget.setItem(row_num, col_num,
                                             QTableWidgetItem(data_for_table.iloc[row_num, col_num]))

        self.retranslateUi(Form)

        QMetaObject.connectSlotsByName(Form)
    # setupUi

    def retranslateUi(self, Form):
        Form.setWindowTitle(QCoreApplication.translate("Form", u"\u0412\u044b\u043f\u0443\u0441\u043a \u041e\u0442\u0447\u0451\u0442\u0430", None))
        self.label.setText(QCoreApplication.translate("Form", u"\u0410\u0434\u0440\u0435\u0441:", None))
        self.label_2.setText(QCoreApplication.translate("Form", u"\u0412\u043b\u0430\u0434\u0435\u043b\u0435\u0446:", None))
        self.label_3.setText(QCoreApplication.translate("Form", u"\u0413\u043e\u0440\u043e\u0434:", None))
        ___qtablewidgetitem = self.tableWidget.horizontalHeaderItem(0)
        ___qtablewidgetitem.setText("ПК");
        ___qtablewidgetitem1 = self.tableWidget.horizontalHeaderItem(1)
        ___qtablewidgetitem1.setText("Вид");
        ___qtablewidgetitem2 = self.tableWidget.horizontalHeaderItem(2)
        ___qtablewidgetitem2.setText("Кол-во Факт.");
        ___qtablewidgetitem3 = self.tableWidget.horizontalHeaderItem(3)
        ___qtablewidgetitem3.setText("Кол-во Адм.");
        ___qtablewidgetitem4 = self.tableWidget.horizontalHeaderItem(4)
        ___qtablewidgetitem4.setText("Кол-во Пред.");
        ___qtablewidgetitem5 = self.tableWidget.horizontalHeaderItem(5)
        ___qtablewidgetitem5.setText("Условия содержания")

        #___qtablewidgetitem2 = self.tableWidget.horizontalHeaderItem(2) #разкомменть и две станут колво факт, походу где то одно ссылается на другое
        #___qtablewidgetitem2.setText("Кол-во Факт.");

        self.label_4.setText(QCoreApplication.translate("Form", u"\u0416\u0438\u0432\u043e\u0442\u043d\u044b\u0435 \u0445\u043e\u0437\u044f\u0439\u0441\u0442\u0432\u0430", None))
        self.documentAddAnimal.setText(QCoreApplication.translate("Form", u"\u0414\u043e\u0431\u0430\u0432\u0438\u0442\u044c \u0416\u0438\u0432\u043e\u0442\u043d\u043e\u0435", None))
#if QT_CONFIG(tooltip)
#         self.createHouseholdReport.setToolTip(QCoreApplication.translate("Form", u"\u0412\u044b\u043f\u0443\u0441\u043a \u043e\u0442\u0447\u0451\u0442\u0430 \u0438 \u0437\u0430\u043f\u0438\u0441\u044c \u0432 \u0444\u0430\u0439\u043b", u"\u0412\u044b\u043f\u0443\u0441\u043a \u043e\u0442\u0447\u0451\u0442\u0430 \u0438 \u0437\u0430\u043f\u0438\u0441\u044c \u0432 \u0444\u0430\u0439\u043b"))
# #endif // QT_CONFIG(tooltip)
#         self.createHouseholdReport.setText(QCoreApplication.translate("Form", u"\u0412\u044b\u043f\u0443\u0441\u0442\u0438\u0442\u044c \u0431\u0435\u0448\u0435\u043d\u044b\u0445 \u043f\u0438\u0434\u043e\u0440\u043e\u0432", None))
        self.documentDeleteAnimal.setText(QCoreApplication.translate("Form", u"\u0423\u0434\u0430\u043b\u0438\u0442\u044c \u0416\u0438\u0432\u043e\u0442\u043d\u043e\u0435", None))
        self.documentAddAnimal_2.setText(QCoreApplication.translate("Form", u"\u0420\u0435\u0434\u0430\u043a\u0442\u0438\u0440\u043e\u0432\u0430\u0442\u044c \u0416\u0438\u0432\u043e\u0442\u043d\u043e\u0435", None))
    # retranslateUi

